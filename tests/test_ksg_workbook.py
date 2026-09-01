from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_io import ExcelParseError, load_ksg_excel, load_ksg_workbook


def _write_ksg_header(ws) -> None:
    ws.append(
        [
            "№ талона",
            "Врач",
            "Код услуги",
            "Сумма к оплате",
            "Дата рождения",
            "КСЛП итоговый",
            "КЗ",
            "Поступление",
            "Выписка",
        ]
    )
    ws.append(
        [
            "1",
            "Иванов И.И.",
            "A16.08.001",
            "50000",
            "01.01.1980",
            "0",
            "1.5",
            "01.06.2026",
            "05.06.2026",
        ]
    )


def _write_legacy_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    _write_ksg_header(ws)
    wb.save(path)


def _write_registry_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "КСГ"
    _write_ksg_header(ws)
    other = wb.create_sheet("Др. услуги")
    other.append(["Свод по другим услугам"])
    other.append(["№ талона", "Врач", "Сумма к оплате"])
    other.append(["2", "Петров П.П.", "12000"])
    wb.save(path)


def test_load_ksg_workbook_legacy_single_sheet(tmp_path):
    path = tmp_path / "june.xlsx"
    _write_legacy_workbook(path)

    loaded = load_ksg_workbook(path)

    assert len(loaded.ksg.dataframe) == 1
    assert loaded.ksg.sheet_name == "Лист1"
    assert loaded.other_services is None


def test_load_ksg_excel_prefers_named_sheet_then_fallback(tmp_path):
    path = tmp_path / "fallback.xlsx"
    _write_legacy_workbook(path)

    loaded = load_ksg_excel(path, preferred_sheets=("КСГ",))

    assert loaded.sheet_name == "Лист1"
    assert loaded.dataframe.iloc[0]["№ талона"] == "1"


def test_load_ksg_excel_strict_other_sheet_without_ksg_headers(tmp_path):
    path = tmp_path / "registry.xlsx"
    _write_registry_workbook(path)

    with pytest.raises(ExcelParseError):
        load_ksg_excel(
            path,
            preferred_sheets=("Др. услуги",),
            strict_preferred_sheets=True,
        )


def test_load_ksg_workbook_skips_other_when_headers_do_not_match(tmp_path):
    path = tmp_path / "registry.xlsx"
    _write_registry_workbook(path)

    loaded = load_ksg_workbook(path)

    assert loaded.ksg.sheet_name == "КСГ"
    assert len(loaded.ksg.dataframe) == 1
    assert loaded.other_services is None


def test_load_ksg_workbook_strict_missing_other_sheet_name(tmp_path):
    path = tmp_path / "only_ksg.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "КСГ"
    _write_ksg_header(ws)
    wb.save(path)

    loaded = load_ksg_workbook(path)

    assert loaded.other_services is None


@pytest.mark.skipif(
    not Path("Ксг за июнь.xlsx").exists(),
    reason="нет помесячного КСГ (старый формат)",
)
def test_load_legacy_single_sheet_ksg_workbook():
    path = "Ксг за июнь.xlsx"
    loaded = load_ksg_workbook(path)
    assert len(loaded.ksg.dataframe) > 0
    assert loaded.ksg.sheet_name == "Лист1"
    assert loaded.other_services is None


@pytest.mark.skipif(
    not Path("Стационар_Реестр учета МП на основе КСГ 2.xlsx").exists(),
    reason="нет общего реестра КСГ",
)
def test_load_hospital_ksg_workbook():
    path = Path("Стационар_Реестр учета МП на основе КСГ 2.xlsx")
    loaded = load_ksg_workbook(path)
    assert len(loaded.ksg.dataframe) > 0
    assert "Отделение" in loaded.ksg.dataframe.columns
    if loaded.other_services is not None:
        assert len(loaded.other_services.dataframe) >= 0
        assert loaded.other_services.sheet_name != loaded.ksg.sheet_name
